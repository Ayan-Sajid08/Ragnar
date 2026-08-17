from io import BytesIO

from openpyxl import load_workbook

from services.chunker import chunk_text


def extract_text(file_bytes: bytes):
    workbook = load_workbook(
        BytesIO(file_bytes),
        data_only=True,
    )

    text_chunks = []
    chunk_index = 0

    for worksheet in workbook.worksheets:
        rows = []

        for row in worksheet.iter_rows(values_only=True):
            values = []

            for value in row:
                if value is not None:
                    values.append(str(value).strip())

            if values:
                rows.append(" | ".join(values))

        if not rows:
            continue

        text = (
            f"Sheet: {worksheet.title}\n"
            + "\n".join(rows)
        )

        chunks = chunk_text(
            text=text,
            page_number=0,
            start_index=chunk_index,
        )

        text_chunks.extend(chunks)
        chunk_index += len(chunks)

    return text_chunks, False